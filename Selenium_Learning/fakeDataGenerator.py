#https://faker.readthedocs.io/en/master/

from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake = Faker()

# for i in range(1,11):
#     for j in range(1,4):
#         ws.cell(row = i,column = 1).value=fake.name()
#         ws.cell(row = i,column = 2).value=fake.email()
#         ws.cell(row = i,column = 3).value=fake.address()

#wb.save("fake.xlsx")

print(fake.phone_number())
