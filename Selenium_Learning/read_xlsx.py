from openpyxl import Workbook,load_workbook

wb = load_workbook(filename="first.xlsx")

sh = wb["Sheet"]

row_ct = sh.max_row
col_ct = sh.max_column

for i in range(1,row_ct+1):
    for j in range(1,col_ct+1):
        print(sh.cell(i,j).value,end=" ")
    print()

# print(sh["A2"].value)
# print(wb["Sheet"]["A2"].value)
# print(sh.cell(2,2).value)